"""
Export service for generating Excel and PDF schedules.
"""
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER


class ExportService:
    """Service for exporting schedules to Excel and PDF formats."""

    def __init__(self):
        self.days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        self.time_slots = ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00',
                          '14:00', '15:00', '16:00', '17:00', '18:00', '19:00', '20:00']

    def generate_excel_schedule(self, schedule_data: List[Dict[str, Any]],
                                semester_name: str, department_name: str) -> bytes:
        """
        Generate a professional university-style Excel schedule.

        Args:
            schedule_data: List of schedule entries
            semester_name: Name of the semester
            department_name: Name of the department

        Returns:
            Excel file as bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"

        # Header styling
        header_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        subheader_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        subheader_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell_font = Font(name='Arial', size=10)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # University header
        ws.merge_cells('A1:I1')
        ws['A1'] = "HITU UNIVERSITY"
        ws['A1'].font = Font(name='Arial', size=20, bold=True, color='1F4E78')
        ws['A1'].alignment = header_alignment

        ws.merge_cells('A2:I2')
        ws['A2'] = f"{department_name} - {semester_name}"
        ws['A2'].font = Font(name='Arial', size=14, bold=True, color='4472C4')
        ws['A2'].alignment = header_alignment

        ws.merge_cells('A3:I3')
        ws['A3'] = f"Generated on: {datetime.now().strftime('%B %d, %Y')}"
        ws['A3'].font = Font(name='Arial', size=10, italic=True, color='666666')
        ws['A3'].alignment = header_alignment

        # Schedule grid header
        ws['A5'] = "Time"
        ws['A5'].font = header_font
        ws['A5'].fill = header_fill
        ws['A5'].alignment = header_alignment

        for idx, day in enumerate(self.days, start=2):
            cell = ws.cell(row=5, column=idx)
            cell.value = day
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Time slots
        for row_idx, time in enumerate(self.time_slots, start=6):
            cell = ws.cell(row=row_idx, column=1)
            cell.value = time
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = cell_alignment
            cell.border = border

        # Populate schedule
        schedule_map = {}
        for entry in schedule_data:
            day_idx = self.days.index(entry['day']) if entry['day'] in self.days else 0
            time_idx = self.time_slots.index(entry['start_time']) if entry['start_time'] in self.time_slots else 0
            row = 6 + time_idx
            col = 2 + day_idx
            schedule_map[(row, col)] = entry

        for (row, col), entry in schedule_map.items():
            cell = ws.cell(row=row, column=col)
            cell_value = f"{entry['course_code']}\n{entry['course_name']}\n"
            cell_value += f"Hall: {entry['hall_name']}\n"
            cell_value += f"Dr. {entry['doctor_name']}"
            cell.value = cell_value
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border

            # Color code by department
            dept_colors = {
                'CS': 'D9E1F2',
                'IT': 'F2DCDB',
                'IS': 'E2EFDA',
                'SE': 'FFF2CC'
            }
            dept_code = entry.get('department_code', 'CS')
            fill_color = dept_colors.get(dept_code, 'F2F2F2')
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        # Column widths
        ws.column_dimensions['A'].width = 8
        for col in range(2, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Row heights
        for row in range(6, 19):
            ws.row_dimensions[row].height = 60

        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_pdf_schedule(self, schedule_data: List[Dict[str, Any]],
                            semester_name: str, department_name: str) -> bytes:
        """
        Generate a landscape PDF schedule with university branding.

        Args:
            schedule_data: List of schedule entries
            semester_name: Name of the semester
            department_name: Name of the department

        Returns:
            PDF file as bytes
        """
        from io import BytesIO
        output = BytesIO()

        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1F4E78'),
            alignment=TA_CENTER,
            spaceAfter=12
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4472C4'),
            alignment=TA_CENTER,
            spaceAfter=12
        )

        date_style = ParagraphStyle(
            'CustomDate',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            alignment=TA_CENTER,
            italic=True
        )

        elements = []

        # Header
        elements.append(Paragraph("HITU UNIVERSITY", title_style))
        elements.append(Paragraph(f"{department_name} - {semester_name}", subtitle_style))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y')}", date_style))
        elements.append(Spacer(1, 20))

        # Schedule table
        table_data = [['Time'] + self.days]

        for time in self.time_slots:
            row = [time]
            for day in self.days:
                # Find schedule entry for this day and time
                entry = next(
                    (e for e in schedule_data if e['day'] == day and e['start_time'] == time),
                    None
                )
                if entry:
                    cell_text = f"{entry['course_code']}<br/>{entry['course_name']}<br/>"
                    cell_text += f"Hall: {entry['hall_name']}<br/>Dr. {entry['doctor_name']}"
                    row.append(cell_text)
                else:
                    row.append('')
            table_data.append(row)

        table = Table(table_data, colWidths=[0.8*inch] + [1.5*inch]*7)

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 1), (0, -1), colors.whitesmoke),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWHEIGHTS', (0, 0), (-1, -1), 0.6*inch),
        ])

        table.setStyle(table_style)
        elements.append(table)

        doc.build(elements)
        output.seek(0)
        return output.getvalue()

    def generate_excel_analytics(self, analytics_data: Dict[str, Any],
                                report_name: str) -> bytes:
        """
        Generate Excel analytics report.

        Args:
            analytics_data: Dictionary containing analytics data
            report_name: Name of the report

        Returns:
            Excel file as bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analytics"

        from io import BytesIO
        output = BytesIO()

        # Header
        ws.merge_cells('A1:D1')
        ws['A1'] = f"HITU UNIVERSITY - {report_name}"
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='1F4E78')
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:D2')
        ws['A2'] = f"Generated on: {datetime.now().strftime('%B %d, %Y')}"
        ws['A2'].font = Font(name='Arial', size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Summary section
        row = 4
        ws[f'A{row}'] = "Summary"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
        row += 1

        for key, value in analytics_data.get('summary', {}).items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
            ws[f'B{row}'].font = Font(name='Arial', size=11)
            row += 1

        row += 1

        # Detailed data
        ws[f'A{row}'] = "Details"
        ws[f'A{row}'].font = Font(name='Arial', size=14, bold=True, color='1F4E78')
        row += 1

        # Headers
        headers = analytics_data.get('headers', [])
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # Data rows
        for data_row in analytics_data.get('data', []):
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='center')
            row += 1

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output)
        output.seek(0)
        return output.getvalue()
